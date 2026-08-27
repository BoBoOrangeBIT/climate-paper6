#!/usr/bin/env python3
"""Build the GLACE Phase A records template for one country x ac_type cell.

Column order follows GLACE_extraction_prompt_KEN_v1 Section 5 exactly
(5.1 identification/sampling, 5.2 price, 5.3 efficiency, 5.4 extension),
plus the gap_codes / qc_flag / notes columns the tracing prompt expects
in the Phase A file. The tracing/final layers appended in Phase B are
documented on their own sheet in append order.

Usage: python3 make_records_template.py OUT.xlsx
"""
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

PHASE_A_COLUMNS = [
    # 5.1 identification and sampling
    ("country", "KEN", "parameter", "yes"),
    ("platform", "platform name", "parameter", "yes"),
    ("platform_type", "marketplace / vertical_dealer", "parameter", "yes"),
    ("delivery_city", "Nairobi 00100", "Step 0", "yes"),
    ("rank", "integer, consecutive from 1, qualifying records only", "search page", "yes"),
    ("is_sponsored", "true / false", "search page", "yes"),
    ("frame_position", "position in raw search results incl. failed listings", "search page", "recommended"),
    ("brand", "brand name", "detail page", "yes"),
    ("manufacturer_mpn", "manufacturer model code, verbatim casing and hyphens", "spec table", "yes"),
    ("mpn_missing", "true / false", "derived", "yes"),
    ("platform_sku", "platform item id, taken from the URL", "URL", "yes"),
    ("ac_type", "split", "parameter", "yes"),
    ("listing_condition", "fixed at new", "detail page", "yes"),
    ("scope_status", "in_scope / out_of_scope_window / out_of_scope_portable / out_of_scope_central / out_of_scope_vrf", "derived", "yes"),
    ("capture_date", "YYYY-MM-DD in UTC", "system", "yes"),
    ("snapshot_id", "snapshot file name", "Step 4", "yes"),
    ("snapshot_sha256", "64 hex characters", "Step 4", "yes"),
    # 5.2 price
    ("price_list", "struck-through/original price, numeric, no thousands separator", "detail page", "if available"),
    ("price_current", "current selling price, numeric", "detail page", "yes"),
    ("price_list_label", "verbatim label text of the struck-through price", "detail page", "if available"),
    ("currency", "KES", "parameter", "yes"),
    ("price_basis", "single_merchant_list_price", "parameter", "yes"),
    ("price_tax_status", "vat_inclusive_16pct default; vat_exclusive if page states pre-tax", "detail page", "yes"),
    ("installation_bundled", "true / false / unstated", "detail page", "yes"),
    ("price_source_url", "full product detail page URL", "URL", "yes"),
    ("price_flag", "blank or unresolved when price gap ratio > 0.3", "derived", "conditional"),
    # 5.3 efficiency
    ("capacity_raw_value", "capacity figure exactly as printed", "spec table", "yes"),
    ("capacity_raw_unit", "BTU/h mostly; kW verbatim if shown", "spec table", "yes"),
    ("rated_cooling_capacity_w", "numeric, watts, converted at 0.29307 W per BTU/h", "derived", "yes"),
    ("rated_power_input_w", "numeric, watts", "spec table", "yes"),
    ("rated_eer", "capacity / input power, two decimals", "computed", "yes"),
    ("capacity_source", "retail_page at this stage", "derived", "yes"),
    ("efficiency_retail_claimed", "efficiency value stated by the retailer", "detail page", "if available"),
    ("efficiency_metric_claimed", "metric name exactly as stated", "detail page", "if available"),
    ("label_grade_reported", "star count or class letter, verbatim", "detail page or label", "if available"),
    ("label_version", "label edition, year or standard part", "label image", "yes where a grade is shown"),
    ("efficiency_source_type", "label_image_official / retail_page_claim / none", "derived", "yes"),
    ("efficiency_source_url", "URL of label image or spec table; detail page URL if none", "detail page", "yes"),
    ("match_ambiguous", "true / false", "derived", "yes"),
    # 5.4 extension
    ("installation_label", "verbatim text", "detail page", "if shown"),
    ("installation_price", "numeric, KES, only when quoted separately", "detail page", "if shown"),
    ("refrigerant", "R22 / R410A / R32 / R290 / R454B", "detail page", "if shown"),
    ("compressor_type", "fixed / inverter", "detail page", "if shown"),
    ("reversible_heat_pump", "true / false (only from a listed rated heating output)", "detail page", "if shown"),
    ("noise_level_db", "numeric", "detail page", "if shown"),
    ("smart_features", "true / false", "detail page", "if shown"),
    # carried in the Phase A file per the tracing prompt's input definition
    ("gap_codes", "A / B / C / C-geo / D / X per blank field", "derived", "every blank field"),
    ("qc_flag", "blank / eer_out_of_range / unresolved_match", "derived", "conditional"),
    ("notes", "free text", "operator", "optional"),
]

# Appended in Phase B (tracing prompt Section 5), in this order.
TRACING_FINAL_COLUMNS = [
    ("efficiency_source_tier", "t1_registry / t1_manufacturer_spec / t1c_label_image / t2_retail_claim / none", "tracing layer"),
    ("mpn_recovered_from", "title / spec_table / label_image / none; only where mpn_missing true", "tracing layer"),
    ("tracing_priority", "1 or 2, per the Checkpoint 1 worklist", "tracing layer"),
    ("registry_name", "register that produced the hit", "tracing layer"),
    ("registry_record_id", "record identifier in the registry", "tracing layer"),
    ("registry_match_type", "exact_mpn / wildcard_mpn / brand_model / none", "tracing layer"),
    ("match_candidate_count", "integer", "tracing layer"),
    ("efficiency_native", "native efficiency value from the authoritative source (EER for KEN)", "tracing layer"),
    ("efficiency_metric", "native metric name, verbatim", "tracing layer"),
    ("label_grade_auth", "grade recorded by the authoritative source, verbatim", "tracing layer"),
    ("label_version_auth", "label edition, year or standard part; none if unstated", "tracing layer"),
    ("capacity_raw_value_auth", "capacity figure as printed by the source", "tracing layer"),
    ("capacity_raw_unit_auth", "unit as printed by the source", "tracing layer"),
    ("rated_cooling_capacity_w_auth", "numeric, watts (0.29307 W per BTU/h)", "tracing layer"),
    ("rated_power_input_w_auth", "numeric, watts", "tracing layer"),
    ("snapshot_auth_id", "snapshot file name", "tracing layer"),
    ("snapshot_auth_sha256", "64 hex characters", "tracing layer"),
    ("rated_cooling_capacity_w_retail", "Phase A retail value, moved here before adjudication overwrites the unsuffixed column", "final layer"),
    ("rated_power_input_w_retail", "Phase A retail value, preserved", "final layer"),
    ("rated_eer_retail", "Phase A retail value, preserved", "final layer"),
    ("capacity_source_retail", "Phase A value (retail_page), preserved", "final layer"),
    ("capacity_deviation_pct", "(retail - authoritative) / authoritative x 100, two decimals", "final layer"),
    ("eer_deviation_pct", "same basis", "final layer"),
]

NOTE_ROWS = [
    "Phase B rules: efficiency_source_url is repointed at the authoritative source (retail URL stays in price_source_url).",
    "Adjudicated values are written into the existing unsuffixed columns rated_cooling_capacity_w, rated_power_input_w, rated_eer, capacity_source; the retail originals move to the _retail columns above. qc_flag and match_ambiguous are updated in place.",
    "qc_flag values after Phase B: blank / eer_out_of_range / large_deviation / metric_mismatch / unresolved_match.",
]


def build(path):
    wb = Workbook()
    header_font = Font(name="Arial", bold=True, size=10)
    body_font = Font(name="Arial", size=10)
    example_fill = PatternFill("solid", fgColor="FFF2CC")
    layer_fill = PatternFill("solid", fgColor="DDEBF7")

    ws = wb.active
    ws.title = "records"
    for col, (name, _, _, _) in enumerate(PHASE_A_COLUMNS, 1):
        c = ws.cell(row=1, column=col, value=name)
        c.font = header_font
        ws.column_dimensions[c.column_letter].width = max(14, min(28, len(name) + 2))
    example = {
        "country": "KEN", "platform": "Jumia Kenya", "platform_type": "marketplace",
        "delivery_city": "Nairobi 00100", "rank": 1, "is_sponsored": "false",
        "frame_position": 1, "brand": "LG", "manufacturer_mpn": "S4-Q12JA3QD",
        "mpn_missing": "false", "platform_sku": "GE1128EAB8VJ2NAFAMZ", "ac_type": "split",
        "listing_condition": "new", "scope_status": "in_scope", "capture_date": "2026-08-27",
        "snapshot_id": "KEN_split_jumia_1_GE1128EAB8VJ2NAFAMZ_20260827T083000Z.png",
        "snapshot_sha256": "0" * 64,
        "price_list": 115995, "price_current": 84999, "price_list_label": "KSh 115,995",
        "currency": "KES", "price_basis": "single_merchant_list_price",
        "price_tax_status": "vat_inclusive_16pct", "installation_bundled": "unstated",
        "price_source_url": "https://www.jumia.co.ke/example-product.html",
        "price_flag": "unresolved",
        "capacity_raw_value": 12000, "capacity_raw_unit": "BTU/h",
        "rated_cooling_capacity_w": 3516.84, "rated_power_input_w": 1100,
        "capacity_source": "retail_page", "efficiency_retail_claimed": 3.2,
        "efficiency_metric_claimed": "EER", "label_grade_reported": "3",
        "label_version": "EPRA 2016", "efficiency_source_type": "retail_page_claim",
        "efficiency_source_url": "https://www.jumia.co.ke/example-product.html",
        "match_ambiguous": "false",
        "refrigerant": "R410A", "compressor_type": "inverter", "smart_features": "false",
        "gap_codes": "B(price_list_label-example)", "notes": "EXAMPLE ROW - delete before the run",
    }
    names = [c[0] for c in PHASE_A_COLUMNS]
    for col, name in enumerate(names, 1):
        c = ws.cell(row=2, column=col, value=example.get(name, ""))
        c.font = body_font
        c.fill = example_fill
    eer_col = names.index("rated_eer") + 1
    # value, not a live formula: the run records computed values (3516.84 / 1100, 2 dp)
    c = ws.cell(row=2, column=eer_col, value=3.20)
    c.fill = example_fill
    c.font = body_font
    ws.freeze_panes = "A2"

    wd = wb.create_sheet("tracing_final_append")
    for col, h in enumerate(["column (append order)", "values and format", "layer"], 1):
        wd.cell(row=1, column=col, value=h).font = header_font
    for r, (name, fmt, layer) in enumerate(TRACING_FINAL_COLUMNS, 2):
        wd.cell(row=r, column=1, value=name).font = body_font
        wd.cell(row=r, column=2, value=fmt).font = body_font
        c = wd.cell(row=r, column=3, value=layer)
        c.font = body_font
        c.fill = layer_fill
    r = len(TRACING_FINAL_COLUMNS) + 3
    for note in NOTE_ROWS:
        wd.cell(row=r, column=1, value=note).font = body_font
        r += 1
    wd.column_dimensions["A"].width = 34
    wd.column_dimensions["B"].width = 80
    wd.column_dimensions["C"].width = 14

    wf = wb.create_sheet("field_dictionary")
    for col, h in enumerate(["field", "values and format", "source", "required"], 1):
        wf.cell(row=1, column=col, value=h).font = header_font
    for r, (name, fmt, src, req) in enumerate(PHASE_A_COLUMNS, 2):
        for col, v in enumerate([name, fmt, src, req], 1):
            wf.cell(row=r, column=col, value=v).font = body_font
    wf.column_dimensions["A"].width = 30
    wf.column_dimensions["B"].width = 80
    wf.column_dimensions["C"].width = 20
    wf.column_dimensions["D"].width = 22

    wl = wb.create_sheet("legend")
    lines = [
        "GLACE records template - KEN x split - RUN_ID KEN_split_YYYYMMDD",
        "Sheet 'records': Phase A output, one row per qualifying listing, columns in the exact Section 5 order.",
        "Row 2 (yellow) is an EXAMPLE with realistic formats only - delete it before the run. Every other cell is filled by the extraction run; there are no precomputed values.",
        "rated_eer is the only permitted computation (example shows the formula); rated_cooling_capacity_w uses the authorised 0.29307 W per BTU/h conversion only.",
        "Every blank field carries a gap code in gap_codes (A/B/C/C-geo/D/X). Never impute, never delete a record.",
        "Sheet 'tracing_final_append': columns Phase B appends, in order, with the adjudication rules.",
        "Sheet 'field_dictionary': per-field format, source and requirement, transcribed from the prompt.",
    ]
    for r, line in enumerate(lines, 1):
        c = wl.cell(row=r, column=1, value=line)
        c.font = body_font
        c.alignment = Alignment(wrap_text=True)
    wl.column_dimensions["A"].width = 120

    wb.save(path)
    print(f"wrote {path}")


if __name__ == "__main__":
    build(sys.argv[1] if len(sys.argv) > 1 else "records_template.xlsx")
